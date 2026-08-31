"""
Report Export Utility Module.
Generates styled Excel (.xlsx) and standard CSV (.csv) reports using OpenPyXL and Pandas.
"""

import io
from datetime import datetime
from typing import Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.database import get_export_dataframe


def generate_csv_report(only_approved: bool = True) -> bytes:
    """Generate CSV byte stream of certificate records."""
    df = get_export_dataframe(only_approved=only_approved)
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False)
    return csv_buffer.getvalue().encode("utf-8")


def generate_excel_report(only_approved: bool = True) -> bytes:
    """
    Generate professional, styled Excel (.xlsx) workbook using OpenPyXL.
    Includes styled headers, status highlights, auto column widths, and summary stats.
    """
    df = get_export_dataframe(only_approved=only_approved)

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    title_suffix = "Approved Records" if only_approved else "All Records"
    ws.title = f"KTU {title_suffix}"

    # Palette Styles
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="4B5563")
    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Status Fills
    status_styles = {
        "Approved": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Recommended": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "Manual Verification Required": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Flagged": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Rejected": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    }

    # 1. Report Title Header Block
    ws.merge_cells("A1:J1")
    ws["A1"] = "APJ Abdul Kalam Technological University (KTU)"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:J2")
    report_desc = f"Student Activity Point Verification Report — {title_suffix} (Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')})"
    ws["A2"] = report_desc
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10  # blank space

    # 2. Table Column Headers (Row 4)
    start_row = 4
    headers = list(df.columns)

    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[start_row].height = 26

    # 3. Data Rows
    current_row = start_row + 1
    total_awarded_points = 0

    for idx, row in df.iterrows():
        for col_idx, col_name in enumerate(headers, 1):
            val = row[col_name]
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border

            # Row zebra striping
            if idx % 2 == 1:
                cell.fill = zebra_fill

            # Alignment heuristics
            if col_name in ["Certificate ID", "Register Number", "Semester", "Confidence (%)", "Suggested Points", "Awarded Points", "Status", "Event Date", "Verification Date", "Submission Date"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Status cell color highlighting
            if col_name == "Status" and str(val) in status_styles:
                cell.fill = status_styles[str(val)]
                cell.font = bold_font

            # Sum up points
            if col_name == "Awarded Points" and pd.notnull(val):
                try:
                    total_awarded_points += int(val)
                except Exception:
                    pass

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # 4. Summary Totals Footer Row
    if not df.empty:
        summary_row = current_row
        ws.cell(row=summary_row, column=1, value="TOTAL")
        ws.cell(row=summary_row, column=1).font = bold_font
        ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=summary_row, column=1).border = thin_border

        # Merge from col 1 to col right before 'Awarded Points'
        try:
            pts_col_idx = headers.index("Awarded Points") + 1
            ws.cell(row=summary_row, column=pts_col_idx, value=total_awarded_points)
            ws.cell(row=summary_row, column=pts_col_idx).font = bold_font
            ws.cell(row=summary_row, column=pts_col_idx).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=summary_row, column=pts_col_idx).fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            ws.cell(row=summary_row, column=pts_col_idx).border = thin_border
        except ValueError:
            pass

        ws.row_dimensions[summary_row].height = 22

    # 5. Auto-fit column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            # Skip title merged row for width calculation
            if cell.row < start_row:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to byte buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()
